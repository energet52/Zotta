"""Multi-format GL export service.

Supports CSV, Excel, PDF, JSON, and XML export for:
- Journal entries (filtered)
- Trial balance
- Account ledger
- Financial statements
- Custom reports

All exports are logged for audit purposes.
"""

import csv
import io
import json
import logging
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any
from xml.etree import ElementTree as ET

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.gl import GLExportLog

logger = logging.getLogger(__name__)


class DecimalEncoder(json.JSONEncoder):
    """JSON encoder that handles Decimal types."""
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)


# ---------------------------------------------------------------------------
# Format renderers
# ---------------------------------------------------------------------------

def export_csv(
    data: list[dict],
    columns: list[str] | None = None,
) -> bytes:
    """Render data as CSV bytes."""
    if not data:
        return b""
    cols = columns or list(data[0].keys())
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=cols, extrasaction="ignore")
    writer.writeheader()
    for row in data:
        writer.writerow({k: row.get(k) for k in cols})
    return output.getvalue().encode("utf-8")


def export_json(data: list[dict], metadata: dict | None = None) -> bytes:
    """Render data as JSON with optional metadata envelope."""
    envelope = {
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "record_count": len(data),
        "data": data,
    }
    if metadata:
        envelope["metadata"] = metadata
    return json.dumps(envelope, cls=DecimalEncoder, indent=2).encode("utf-8")


def export_xml(
    data: list[dict],
    root_tag: str = "GLExport",
    row_tag: str = "Entry",
) -> bytes:
    """Render data as XML bytes."""
    root = ET.Element(root_tag)
    root.set("exported_at", datetime.now(timezone.utc).isoformat())
    root.set("record_count", str(len(data)))

    for row in data:
        entry_el = ET.SubElement(root, row_tag)
        for key, value in row.items():
            child = ET.SubElement(entry_el, key)
            child.text = str(value) if value is not None else ""

    return ET.tostring(root, encoding="unicode", xml_declaration=True).encode("utf-8")


def export_excel(
    data: list[dict],
    columns: list[str] | None = None,
    sheet_name: str = "GL Export",
) -> bytes:
    """Render data as Excel (.xlsx) bytes using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, numbers
    except ImportError:
        logger.warning("openpyxl not installed — falling back to CSV")
        return export_csv(data, columns)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    cols = columns or (list(data[0].keys()) if data else [])

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    number_format = '#,##0.00'
    for row_idx, row in enumerate(data, 2):
        for col_idx, col_name in enumerate(cols, 1):
            value = row.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, (int, float, Decimal)):
                cell.number_format = number_format
                cell.alignment = Alignment(horizontal="right")

    # Auto-filter and freeze panes
    if cols:
        ws.auto_filter.ref = f"A1:{chr(64 + len(cols))}{len(data) + 1}"
    ws.freeze_panes = "A2"

    # Auto-width
    for col_idx, col_name in enumerate(cols, 1):
        max_len = max(
            len(str(col_name)),
            *(len(str(row.get(col_name, ""))) for row in data[:100]),
        ) if data else len(str(col_name))
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"].width = min(max_len + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_pdf(
    data: list[dict],
    columns: list[str] | None = None,
    title: str = "GL Export Report",
) -> bytes:
    """Render data as PDF using reportlab."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        logger.warning("reportlab not installed — falling back to CSV")
        return export_csv(data, columns)

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Spacer(1, 0.25 * inch))
    elements.append(Paragraph(
        f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} | "
        f"Records: {len(data)}",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 0.25 * inch))

    if not data:
        elements.append(Paragraph("No data to display.", styles["Normal"]))
        doc.build(elements)
        return output.getvalue()

    cols = columns or list(data[0].keys())

    # Build table data
    table_data = [[c.replace("_", " ").title() for c in cols]]
    for row in data:
        table_data.append([
            f"{row.get(c, ''):,.2f}" if isinstance(row.get(c), (int, float, Decimal)) else str(row.get(c, ""))
            for c in cols
        ])

    # Create table
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (0, 1), (-1, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]))

    elements.append(table)
    doc.build(elements)
    return output.getvalue()


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

EXPORTERS = {
    "csv": export_csv,
    "json": export_json,
    "xml": export_xml,
    "xlsx": export_excel,
    "pdf": export_pdf,
}


async def export_data(
    db: AsyncSession,
    *,
    data: list[dict],
    format: str,
    columns: list[str] | None = None,
    title: str = "GL Export",
    user_id: int,
    export_type: str = "journal_entries",
    filters: dict | None = None,
) -> bytes:
    """Export data in the requested format and log the export."""
    fmt = format.lower()
    if fmt not in EXPORTERS:
        raise ValueError(f"Unsupported format: {fmt}. Supported: {list(EXPORTERS.keys())}")

    exporter = EXPORTERS[fmt]

    if fmt == "pdf":
        result = exporter(data, columns=columns, title=title)
    elif fmt in ("json",):
        result = exporter(data, metadata={"title": title, "filters": filters})
    elif fmt == "xml":
        result = exporter(data)
    else:
        result = exporter(data, columns=columns)

    # Log export
    log = GLExportLog(
        user_id=user_id,
        export_type=export_type,
        format=fmt,
        filters=filters,
        row_count=len(data),
    )
    db.add(log)
    await db.flush()

    logger.info(
        "Exported %d rows as %s (type=%s) for user %d",
        len(data), fmt, export_type, user_id,
    )
    return result


def get_content_type(format: str) -> str:
    """Return the MIME content type for a format."""
    return {
        "csv": "text/csv",
        "json": "application/json",
        "xml": "application/xml",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "pdf": "application/pdf",
    }.get(format.lower(), "application/octet-stream")


def get_file_extension(format: str) -> str:
    """Return the file extension for a format."""
    return {
        "csv": ".csv",
        "json": ".json",
        "xml": ".xml",
        "xlsx": ".xlsx",
        "pdf": ".pdf",
    }.get(format.lower(), ".dat")
